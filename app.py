# app.py
import os
import sqlite3
from flask import Flask, render_template, request, g, redirect, url_for, send_file
from datetime import datetime
from werkzeug.utils import secure_filename
from evaluator import evaluate_pdfs
from utils import extract_max_marks
import pandas as pd
import io
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = "uploads"
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

DATABASE = 'evaluations.db'

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
    return db

def init_db():
    with app.app_context():
        db = get_db()
        db.execute('''CREATE TABLE IF NOT EXISTS evaluations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            roll_no TEXT NOT NULL,
            subject TEXT NOT NULL,
            timestamp TEXT,
            total_marks REAL,
            percentage REAL,
            grade TEXT,
            grade_point REAL,
            student_pdf_path TEXT,
            model_pdf_path TEXT,
            question_pdf_path TEXT,
            credits INTEGER DEFAULT 3)''')

        db.execute('''CREATE TABLE IF NOT EXISTS students (
            roll_no TEXT PRIMARY KEY,
            full_name TEXT,
            department TEXT)''')

        db.execute('''CREATE TABLE IF NOT EXISTS question_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            evaluation_id INTEGER,
            question TEXT,
            student_answer TEXT,
            model_answer TEXT,
            similarity REAL,
            score REAL,
            max_marks REAL,
            FOREIGN KEY(evaluation_id) REFERENCES evaluations(id))''')
        
        db.commit()

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        try:
            roll_no = request.form.get('roll_no').strip()
            subject = request.form.get('subject').strip()
            credits = int(request.form.get('credits', 3))
            full_name = request.form.get('full_name').strip()

            if not roll_no or not subject:
                return render_template("index.html", error="Roll number and subject are required!")
            if any(f not in request.files for f in ["student_pdf", "model_pdf", "question_pdf"]):
                return render_template("index.html", error="Missing file uploads!")

            paths = {}
            for file_type in ['student_pdf', 'model_pdf', 'question_pdf']:
                file = request.files[file_type]
                if file.filename == '':
                    return render_template("index.html", error="No selected file!")
                filename = f"{roll_no}_{subject}_{file_type}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(filename))
                file.save(filepath)
                paths[file_type] = filepath

            max_marks = extract_max_marks(paths['question_pdf'])
            if not max_marks:
                return render_template("index.html", error="Could not extract marks from question paper!")

            results, total_marks = evaluate_pdfs(
                student_pdf=paths['student_pdf'],
                model_pdf=paths['model_pdf'],
                max_marks=max_marks
            )

            total_max_marks = sum(max_marks.values())/2
            percentage = (total_marks / total_max_marks * 100) if total_max_marks else 0

            if percentage >= 90:
                grade, grade_point = 'O', 10.0
            elif percentage >= 80:
                grade, grade_point = 'A+', 9.0
            elif percentage >= 70:
                grade, grade_point = 'A', 8.0
            elif percentage >= 60:
                grade, grade_point = 'B+', 7.0
            elif percentage >= 50:
                grade, grade_point = 'B', 6.0
            elif percentage >= 40:
                grade, grade_point = 'C', 5.0
            else:
                grade, grade_point = 'F', 0.0

            db = get_db()
            db.execute('INSERT OR REPLACE INTO students (roll_no, full_name) VALUES (?, ?)', (roll_no, full_name))

            db.execute('DELETE FROM evaluations WHERE roll_no = ? AND subject = ?', (roll_no, subject))

            eval_id = db.execute('''INSERT INTO evaluations (
                roll_no, subject, timestamp, total_marks, percentage, grade, grade_point,
                student_pdf_path, model_pdf_path, question_pdf_path, credits)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (roll_no, subject, datetime.now().isoformat(), total_marks, percentage,
                 grade, grade_point, paths['student_pdf'], paths['model_pdf'], paths['question_pdf'], credits)).lastrowid

            for result in results:
                db.execute('''INSERT INTO question_results (
                    evaluation_id, question, student_answer, model_answer,
                    similarity, score, max_marks)
                    VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (eval_id, result["question"], result["student_answer"],
                     result["model_answer"], result["similarity"],
                     result["score"], result["max_marks"]))

            db.commit()

            return render_template("result.html",
                roll_no=roll_no,
                subject=subject,
                total_marks=total_marks,
                percentage=round(percentage, 2),
                grade=grade,
                grade_point=grade_point,
                evaluation_date=datetime.now().strftime("%d %b %Y %H:%M"),
                question_results=results)

        except Exception as e:
            app.logger.error(f"Evaluation error: {str(e)}")
            return render_template("index.html", error=f"An error occurred: {str(e)}")

    return render_template("index.html")

@app.route("/reports")
def reports():
    try:
        db = get_db()

        results = db.execute('''
            SELECT e.*, strftime('%d/%m/%Y %H:%M', e.timestamp) as formatted_date, s.full_name
            FROM evaluations e
            LEFT JOIN students s ON e.roll_no = s.roll_no
            ORDER BY e.timestamp DESC
        ''').fetchall()

        gpa_data = db.execute('''
            SELECT e.roll_no, s.full_name,
                SUM(e.grade_point * e.credits) / SUM(e.credits) as gpa,
                COUNT(*) as exams_taken
            FROM evaluations e
            LEFT JOIN students s ON e.roll_no = s.roll_no
            GROUP BY e.roll_no
        ''').fetchall()

        # Check for any failed subjects (grade point 0) and adjust GPA if necessary
        modified_gpa_data = []
        for student in gpa_data:
            evaluations = db.execute('''
                SELECT grade_point FROM evaluations WHERE roll_no = ?
            ''', (student['roll_no'],)).fetchall()
            failed_subjects = any(eval['grade_point'] == 0.0 for eval in evaluations)
            if failed_subjects:
                modified_gpa_data.append({**dict(student), 'gpa': 0.0})
            else:
                modified_gpa_data.append(dict(student))

        subject_stats = db.execute('''
            SELECT subject,
                AVG(percentage) as avg_percentage,
                COUNT(*) as exams_graded,
                AVG(grade_point) as avg_grade_point,
                credits
            FROM evaluations
            GROUP BY subject
        ''').fetchall()

        return render_template("reports.html",
            results=results,
            gpa_data=modified_gpa_data,
            subject_stats=subject_stats)

    except Exception as e:
        app.logger.error(f"Reports error: {str(e)}")
        return render_template("error.html", message="Could not generate reports")

@app.route("/student/<roll_no>")
def student_report(roll_no):
    try:
        db = get_db()

        student = db.execute('''
            SELECT e.*, strftime('%d/%m/%Y %H:%M', e.timestamp) as formatted_date, s.full_name
            FROM evaluations e
            LEFT JOIN students s ON e.roll_no = s.roll_no
            WHERE e.roll_no = ?
            ORDER BY e.timestamp DESC
            LIMIT 1
        ''', (roll_no,)).fetchone()

        if not student:
            return render_template("error.html", message="Student not found")

        evaluations = db.execute('''
            SELECT e.subject, e.total_marks, e.percentage, e.grade, e.grade_point, e.credits,
                strftime('%d/%m/%Y %H:%M', e.timestamp) as formatted_date
            FROM evaluations e
            WHERE e.roll_no = ?
            ORDER BY e.timestamp DESC
        ''', (roll_no,)).fetchall()

        subject_performance = db.execute('''
            SELECT e.subject,
                e.percentage as avg_percentage,
                e.grade_point as avg_grade_point,
                e.credits
            FROM evaluations e
            INNER JOIN (
                SELECT subject, MAX(timestamp) as latest_time
                FROM evaluations
                WHERE roll_no = ?
                GROUP BY subject
            ) latest_eval
            ON e.subject = latest_eval.subject AND e.timestamp = latest_eval.latest_time
            WHERE e.roll_no = ?
        ''', (roll_no, roll_no)).fetchall()

        if subject_performance:
            total_credits = sum(row['credits'] for row in subject_performance)
            weighted_gpa = sum(row['avg_grade_point'] * row['credits'] for row in subject_performance)

            # Check for any failed subjects (grade point 0)
            failed_subjects = any(row['avg_grade_point'] == 0.0 for row in subject_performance)
            if failed_subjects:
                semester_gpa = 0.0  # Semester GPA is 0 if any subject is failed
            elif total_credits > 0:
                semester_gpa = weighted_gpa / total_credits
            else:
                semester_gpa = 0.0
        else:
            semester_gpa = 0.0

        return render_template("student_report.html",
            roll_no=roll_no,
            student=student,
            evaluations=evaluations,
            subject_performance=subject_performance,
            semester_gpa=semester_gpa)

    except Exception as e:
        app.logger.error(f"Student report error: {str(e)}")
        return render_template("error.html", message="Could not generate student report")

@app.route("/download_excel")
def download_excel():
    db = get_db()

    rows = db.execute('''
        SELECT e.roll_no, s.full_name, e.subject, e.grade, e.grade_point, e.credits
        FROM evaluations e
        LEFT JOIN students s ON e.roll_no = s.roll_no
        INNER JOIN (
            SELECT roll_no, subject, MAX(timestamp) as latest_time
            FROM evaluations
            GROUP BY roll_no, subject
        ) latest
        ON e.roll_no = latest.roll_no AND e.subject = latest.subject AND e.timestamp = latest.latest_time
    ''').fetchall()

    student_data = defaultdict(dict)
    for row in rows:
        roll = row['roll_no']
        student_data[roll]["Student Name"] = row['full_name']
        subject = row['subject']
        student_data[roll][f"{subject} Grade"] = row['grade']
        student_data[roll][f"{subject} GPA"] = round(row['grade_point'], 2)
        student_data[roll][f"{subject} Credits"] = row['credits']

    for roll, subjects in student_data.items():
        gpa_values = [v for k, v in subjects.items() if "GPA" in k]
        credit_values = [v for k, v in subjects.items() if "Credits" in k]
        if gpa_values and credit_values:
            weighted_sum = sum(gpa_values[i] * credit_values[i] for i in range(len(gpa_values)))
            total_credits = sum(credit_values)
            # Check for any failed subjects (grade point 0)
            failed_subjects = any(gpa == 0.0 for gpa in gpa_values)
            if failed_subjects:
                student_data[roll]["Final GPA"] = 0.0
            elif total_credits > 0:
                student_data[roll]["Final GPA"] = round(weighted_sum / total_credits, 2)
            else:
                student_data[roll]["Final GPA"] = 0.0
        else:
            student_data[roll]["Final GPA"] = 0.0

    final_rows = []
    for roll, data in student_data.items():
        row = {"Roll No": roll}
        row.update(data)
        final_rows.append(row)

    df = pd.DataFrame(final_rows).fillna("")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Student Report")
    output.seek(0)

    return send_file(output,
        as_attachment=True,
        download_name="student_performance_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/download_student_report/<roll_no>/<subject>")
def download_student_report(roll_no, subject):
    db = get_db()
    eval_data = db.execute('''
        SELECT e.*, s.full_name
        FROM evaluations e
        LEFT JOIN students s ON e.roll_no = s.roll_no
        WHERE e.roll_no = ? AND e.subject = ?
        ORDER BY e.timestamp DESC
        LIMIT 1
    ''', (roll_no, subject)).fetchone()

    if not eval_data:
        return "Evaluation not found", 404

    df = pd.DataFrame([{
        "Roll No": eval_data['roll_no'],
        "Student Name" : eval_data['full_name'],
        "Subject": eval_data['subject'],
        "Total Marks": eval_data['total_marks'],
        "Percentage": eval_data['percentage'],
        "Grade": eval_data['grade'],
        "GPA": eval_data['grade_point'],
        "Evaluation Date": eval_data['timestamp'],
        "Credits": eval_data['credits']
    }])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Student Evaluation")
    output.seek(0)

    filename = f"{roll_no}_{subject}_report.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/download_question_wise/<roll_no>/<subject>")
def download_question_wise(roll_no, subject):
    try:
        db = get_db()
        eval_data = db.execute('''
            SELECT e.id, s.full_name
            FROM evaluations e
            LEFT JOIN students s ON e.roll_no = s.roll_no
            WHERE e.roll_no = ? AND e.subject = ?
            ORDER BY e.timestamp DESC
            LIMIT 1
        ''', (roll_no, subject)).fetchone()

        if not eval_data:
            app.logger.error(f"No evaluation found for {roll_no} in {subject}")
            return render_template("error.html", message="Evaluation not found"), 404

        question_results = db.execute('''
            SELECT question, student_answer, model_answer, similarity, score, max_marks
            FROM question_results
            WHERE evaluation_id = ?
            ORDER BY question
        ''', (eval_data['id'],)).fetchall()

        if not question_results:
            app.logger.error(f"No question results found for evaluation ID {eval_data['id']}")
            return render_template("error.html", message="Question results not found"), 404

        data = {
            "Question": [r['question'] for r in question_results],
            "Student Answer": [r['student_answer'] for r in question_results],
            "Model Answer": [r['model_answer'] for r in question_results],
            "Similarity (%)": [float(r['similarity']) for r in question_results],
            "Marks Awarded": [float(r['score']) for r in question_results],
            "Max Marks": [float(r['max_marks']) for r in question_results]
        }
        df = pd.DataFrame(data)

        total_marks = df['Marks Awarded'].sum()
        total_max = df['Max Marks'].sum()
        percentage = (total_marks / total_max * 100) if total_max else 0

        summary_row = {
            "Question": "TOTAL",
            "Student Answer": "",
            "Model Answer": "",
            "Similarity (%)": f"{percentage:.2f}%",
            "Marks Awarded": total_marks,
            "Max Marks": total_max
        }
        df = pd.concat([df, pd.DataFrame([summary_row])], ignore_index=True)

        output = io.BytesIO()
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="Question Results")
                
                workbook = writer.book
                worksheet = writer.sheets["Question Results"]
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4e54c8", end_color="4e54c8", fill_type="solid")
                header_alignment = Alignment(vertical="top", wrap_text=True)
                header_border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                
                for col_num, value in enumerate(df.columns.values):
                    cell = worksheet.cell(row=1, column=col_num + 1)
                    cell.value = value
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = header_border
                
                cell_alignment = Alignment(vertical="top", wrap_text=True)
                
                for col_num, column in enumerate(df.columns):
                    max_len = max(df[column].astype(str).map(len).max(), len(str(column)))
                    worksheet.column_dimensions[chr(65 + col_num)].width = min(max_len + 2, 50)
                    for row in range(2, len(df) + 1):
                        cell = worksheet.cell(row=row, column=col_num + 1)
                        cell.alignment = cell_alignment
                
                total_font = Font(bold=True, color="FFFFFF")
                total_fill = PatternFill(start_color="8f94fb", end_color="8f94fb", fill_type="solid")
                
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=len(df), column=col_num)
                    cell.font = total_font
                    cell.fill = total_fill

            output.seek(0)
        except Exception as e:
            app.logger.error(f"Excel generation error: {str(e)}")
            return render_template("error.html", message=f"Failed to generate Excel file: {str(e)}"), 500

        filename = f"{secure_filename(roll_no)}_{secure_filename(subject)}_question_wise.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        app.logger.error(f"Question-wise download error: {str(e)}", exc_info=True)
        return render_template("error.html", message=f"Could not generate question-wise report: {str(e)}"), 500

@app.route("/delete_student/<roll_no>", methods=["POST"])
def delete_student(roll_no):
    try:
        db = get_db()
        eval_ids = [row[0] for row in db.execute("SELECT id FROM evaluations WHERE roll_no = ?", (roll_no,)).fetchall()]
        for eval_id in eval_ids:
            db.execute("DELETE FROM question_results WHERE evaluation_id = ?", (eval_id,))
        db.execute("DELETE FROM evaluations WHERE roll_no = ?", (roll_no,))
        db.execute("DELETE FROM students WHERE roll_no = ?", (roll_no,))
        db.commit()
        return redirect(url_for('reports'))
    except Exception as e:
        app.logger.error(f"Delete student error: {str(e)}")
        return render_template("error.html", message="Could not delete student records")

@app.route("/download_semester_report/<roll_no>")
def download_semester_report(roll_no):
    try:
        db = get_db()

        # Get student information
        student = db.execute('''
            SELECT roll_no, full_name FROM students WHERE roll_no = ?
        ''', (roll_no,)).fetchone()

        if not student:
            return render_template("error.html", message="Student not found"), 404

        # Get latest evaluation for each subject
        subject_performance = db.execute('''
            SELECT e.subject,
                e.percentage as avg_percentage,
                e.grade_point as avg_grade_point,
                e.credits
            FROM evaluations e
            INNER JOIN (
                SELECT subject, MAX(timestamp) as latest_time
                FROM evaluations
                WHERE roll_no = ?
                GROUP BY subject
            ) latest_eval
            ON e.subject = latest_eval.subject AND e.timestamp = latest_eval.latest_time
            WHERE e.roll_no = ?
        ''', (roll_no, roll_no)).fetchall()

        if not subject_performance:
            return render_template("error.html", message="No evaluation data found for this student"), 404

        # Calculate semester GPA
        total_credits = sum(row['credits'] for row in subject_performance)
        weighted_gpa = sum(row['avg_grade_point'] * row['credits'] for row in subject_performance)
        
        # Check for any failed subjects (grade point 0)
        failed_subjects = any(row['avg_grade_point'] == 0.0 for row in subject_performance)
        if failed_subjects:
            semester_gpa = 0.0
        elif total_credits > 0:
            semester_gpa = weighted_gpa / total_credits
        else:
            semester_gpa = 0.0

        # Prepare data for Excel
        data = {
            "Subject": [row['subject'] for row in subject_performance],
            "Percentage": [row['avg_percentage'] for row in subject_performance],
            "Grade Point": [row['avg_grade_point'] for row in subject_performance],
            "Credits": [row['credits'] for row in subject_performance]
        }
        df = pd.DataFrame(data)

        # Add summary row
        summary_row = {
            "Subject": "SEMESTER SUMMARY",
            "Percentage": "",
            "Grade Point": semester_gpa,
            "Credits": total_credits
        }
        df = pd.concat([df, pd.DataFrame([summary_row])], ignore_index=True)

        # Create Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Semester Report")
            
            # Apply formatting
            workbook = writer.book
            worksheet = writer.sheets["Semester Report"]
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4e54c8", end_color="4e54c8", fill_type="solid")
            header_alignment = Alignment(vertical="center", horizontal="center")
            
            for col_num, value in enumerate(df.columns.values):
                cell = worksheet.cell(row=1, column=col_num + 1)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Summary row formatting
            summary_font = Font(bold=True, color="FFFFFF")
            summary_fill = PatternFill(start_color="8f94fb", end_color="8f94fb", fill_type="solid")
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=len(df), column=col_num)
                cell.font = summary_font
                cell.fill = summary_fill
            
            # Set column widths
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = adjusted_width

        output.seek(0)
        
        filename = f"{roll_no}_semester_report.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        app.logger.error(f"Semester report download error: {str(e)}", exc_info=True)
        return render_template("error.html", message=f"Could not generate semester report: {str(e)}"), 500

if __name__ == "__main__":
    init_db()
    app.run(debug=True)